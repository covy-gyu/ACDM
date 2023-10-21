import os
import time

import openpyxl

import conf
from conf import *
from util import *


class ExcelWriter:
    def __init__(self, file_name):
        self.file_name = file_name
        self.field_names = [
            "검사일시", 
            "탄순번", 
            "탄약고번호", 
            "DODIC", 
            "로트번호",
            "이미지데이터 파일경로", 
            "변위센서데이터 파일경로",
            "변위센서결과 파일경로", 
            "변위센서결과이미지 파일경로",
            "구성품",
            "이미지별 결함코드", 
            "탄별 결함코드",
            # "종합 결함코드",
            # "신관",
            # "탄체 결함코드",
            # "추진장약",
            # "날개",
            # "추진약멈치",
            "검사관",
        ]

        file_path = rep_sl_join(DIR_PATH['done']['excel'], self.file_name)
        if os.path.exists(file_path):
            self.workbook = openpyxl.load_workbook(filename=file_path)
            return

        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.worksheets[0])

    def add_bomb(self, bomb):
        data = self.generate_data(bomb)
        self.add_list_to_sheet(data, str(bomb.num))

    def add_list_to_sheet(self, data, sheet_name):
        sheet = self.workbook.create_sheet(title=sheet_name)
        sheet.append(self.field_names)

        for row in data:
            sheet.append(row)

    @staticmethod
    def generate_data(bomb):
        rel_path = conf.PATH_IN_EXCEL == 'relative'

        lot_name = bomb.lot.name
        num = bomb.num

        defect = bomb.defect.get_defect()  # dict | keys=['CAM1' ~ 'CAM4']
        # defect_byparts = bomb.defect.get_defect_by_parts()

        sensor_data_filename = rep_sl(bomb.sensor_data_path).split('/')[-1]
        sensor_data_path = rep_sl_join(DIR_PATH['done']['sensor']['data'], sensor_data_filename)  # Path after move
        sensor_data_path = get_abs_path(sensor_data_path)
        sensor_res_path = bomb.sensor_res_path
        if rel_path:
            sensor_data_path = path_abs2rel(sensor_data_path)
            sensor_res_path['html'] = path_abs2rel(sensor_res_path['html'])
            sensor_res_path['image'] = path_abs2rel(sensor_res_path['image'])

        data = []
        part = ['약포(하),추진약멈치(하),날개', '약포(상),추진약멈치(상)', '탄체(하)', '신관, 탄체(상)']
        for i in range(24):
            img_path = bomb.img_path['CAM' + str(i // 6 + 1)][i % 6]
            if rel_path:
                img_path = path_abs2rel(img_path)
            # img_path = img_path.replace("ACDM", "ACDM_DONE")

            row = [
                # "탄약고번호", "검사일시", "탄순번", "DODIC" "로트번호",
                get_created_time(img_path),
                num,
                "11탄약창",
                "KC256",
                lot_name,
                # 이미지 경로(camera; CAM)
                img_path.replace("ACDM", "ACDM_DONE"),
                # 변위센서 데이터, 결과(html), 이미지 경로
                sensor_data_path,
                sensor_res_path['html'],
                sensor_res_path['image'],
                # 구성품
                part[i // 6],
                # 이미지별 결함 코드
                defect['CAM' + str(i // 6 + 1)][i % 6],
                # 탄별 결함코드
                defect['CAM' + str(i // 6 + 1)][6],
                # # 종합 결함코드
                # "",
                # # 신관
                # defect_byparts['head'],
                # 탄체
                # defect_byparts['body'],
                # # 추진장약
                # defect_byparts['powder'],
                # # 날개
                # defect_byparts['wing'],
                # # 추진약멈치
                # defect_byparts['anchor'],
                # 검사관
            ]
            data.append(row)

            # Use only when i = 0
            sensor_data_path = ''
            sensor_res_path['html'] = ''
            sensor_res_path['image'] = ''

            # Use only once
            defect['CAM' + str(i // 6 + 1)][6] = ''

            # defect_byparts['head'] = ''
            # defect_byparts['body'] = ''
            # defect_byparts['powder'] = ''
            # defect_byparts['wing'] = ''
            # defect_byparts['anchor'] = ''

        return data

    def save_excel(self):
        file_path = rep_sl_join(DIR_PATH['done']['excel'], self.file_name)
        try:
            self.workbook.save(file_path)
        except IOError:
            os.makedirs(os.path.dirname(file_path))
            self.workbook.save(file_path)
